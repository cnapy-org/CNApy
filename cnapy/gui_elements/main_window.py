import io
import json
import os
import traceback
from tempfile import TemporaryDirectory
from zipfile import BadZipFile, ZipFile
import pickle
import xml.etree.ElementTree as ET
from cnapy.flux_vector_container import FluxVectorContainer
from cnapy.core_gui import model_optimization_with_exceptions, except_likely_community_model_error, get_last_exception_string, has_community_error_substring
import cobra
from optlang_enumerator.cobra_cnapy import CNApyModel
from optlang_enumerator.mcs_computation import flux_variability_analysis
from optlang.symbolics import Zero
import numpy as np
import cnapy.resources  # Do not delete this import - it seems to be unused but in fact it provides the menu icons
import matplotlib.pyplot as plt
from typing import Any, Dict
import openpyxl

from qtpy.QtCore import QFileInfo, Qt, Slot, QTimer, QSignalBlocker, QSize
from qtpy.QtGui import QColor, QIcon, QKeySequence
from qtpy.QtWidgets import (QAction, QActionGroup, QApplication, QFileDialog, QStyle,
                            QMainWindow, QMessageBox, QToolBar, QShortcut, QStatusBar, QLabel)
from qtpy.QtWebEngineWidgets import QWebEngineView

from cnapy.appdata import AppData, ProjectData, Scenario
from cnapy.gui_elements.about_dialog import AboutDialog
from cnapy.gui_elements.central_widget import CentralWidget, ModelTabIndex
from cnapy.gui_elements.clipboard_calculator import ClipboardCalculator
from cnapy.gui_elements.config_dialog import ConfigDialog
from cnapy.gui_elements.download_dialog import DownloadDialog
from cnapy.gui_elements.config_cobrapy_dialog import ConfigCobrapyDialog
from cnapy.gui_elements.efmtool_dialog import EFMtoolDialog
from cnapy.gui_elements.flux_feasibility_dialog import FluxFeasibilityDialog
from cnapy.gui_elements.map_view import MapView
from cnapy.gui_elements.escher_map_view import EscherMapView
from cnapy.gui_elements.mcs_dialog import MCSDialog
from cnapy.gui_elements.strain_design_dialog import SDDialog, SDComputationViewer, SDViewer, SDComputationThread
from cnapy.gui_elements.plot_space_dialog import PlotSpaceDialog
from cnapy.gui_elements.in_out_flux_dialog import InOutFluxDialog
from cnapy.gui_elements.reactions_list import ReactionListColumn
from cnapy.gui_elements.rename_map_dialog import RenameMapDialog
from cnapy.gui_elements.yield_optimization_dialog import YieldOptimizationDialog
from cnapy.gui_elements.flux_optimization_dialog import FluxOptimizationDialog
from cnapy.gui_elements.configuration_cplex import CplexConfigurationDialog
from cnapy.gui_elements.configuration_gurobi import GurobiConfigurationDialog
from cnapy.gui_elements.thermodynamics_dialog import ThermodynamicAnalysisTypes, ThermodynamicDialog
import cnapy.utils as utils

SBML_suffixes = "*.xml *.sbml *.xml.gz *.sbml.gz *.xml.zip *.sbml.zip"

class MainWindow(QMainWindow):
    """The cnapy main window"""

    def __init__(self, appdata: AppData):
        QMainWindow.__init__(self)
        self.setWindowTitle("cnapy")
        self.appdata = appdata

        # self.heaton_action and self.onoff_action need to be defined before CentralWidget
        self.heaton_action = QAction("Heatmap coloring", self)
        self.heaton_action.setIcon(QIcon(":/icons/heat.png"))
        self.heaton_action.triggered.connect(self.set_heaton)

        self.onoff_action = QAction("On/Off coloring", self)
        self.onoff_action.setIcon(QIcon(":/icons/onoff.png"))
        self.onoff_action.triggered.connect(self.set_onoff)

        self.central_widget = CentralWidget(self)
        self.setCentralWidget(self.central_widget)

        self.menu = self.menuBar()
        self.file_menu = self.menu.addMenu("&Project")

        new_project_action = QAction("&New project", self)
        new_project_action.setShortcut("Ctrl+N")
        self.file_menu.addAction(new_project_action)
        new_project_action.triggered.connect(self.new_project)

        new_project_from_sbml_action = QAction(
            "New project from SBML...", self)
        self.file_menu.addAction(new_project_from_sbml_action)
        new_project_from_sbml_action.triggered.connect(
            self.new_project_from_sbml)

        open_project_action = QAction("&Open project...", self)
        open_project_action.setShortcut("Ctrl+O")
        self.file_menu.addAction(open_project_action)
        open_project_action.triggered.connect(self.open_project_dialog)

        self.recent_cna_menu = self.file_menu.addMenu("Open recent")
        self.recent_cna_actions: Dict[str, QAction] = {}

        self.save_project_action = QAction("&Save project", self)
        self.save_project_action.setShortcut("Ctrl+S")
        self.file_menu.addAction(self.save_project_action)
        self.save_project_action.triggered.connect(self.save_project)

        save_as_project_action = QAction("&Save project as...", self)
        save_as_project_action.setShortcut("Ctrl+Shift+S")
        self.file_menu.addAction(save_as_project_action)
        save_as_project_action.triggered.connect(self.save_project_as)

        export_sbml_action = QAction("Export SBML...", self)
        self.file_menu.addAction(export_sbml_action)
        export_sbml_action.triggered.connect(self.export_sbml)

        download_examples = QAction("Download CNApy example projects...", self)
        self.file_menu.addAction(download_examples)
        download_examples.triggered.connect(self.download_examples)

        exit_action = QAction("Exit", self)
        exit_action.setShortcut("Ctrl+Q")
        self.file_menu.addAction(exit_action)
        exit_action.triggered.connect(self.exit_app)

        self.scenario_menu = self.menu.addMenu("Scenario")

        load_default_scenario_action = QAction("Apply default scenario flux values", self)
        self.scenario_menu.addAction(load_default_scenario_action)
        load_default_scenario_action.setIcon(QIcon(":/icons/d-font.png"))
        load_default_scenario_action.triggered.connect(
            self.load_default_scenario)

        set_scenario_to_default_scenario_action = QAction(
            "Set current scenario fluxes as default scenario fluxes", self)
        self.scenario_menu.addAction(set_scenario_to_default_scenario_action)
        set_scenario_to_default_scenario_action.triggered.connect(
            self.set_scenario_to_default_scenario)

        load_scenario_action = QAction("Load scenario...", self)
        load_scenario_action.setIcon(self.style().standardIcon(QStyle.SP_DirIcon))
        self.scenario_menu.addAction(load_scenario_action)
        load_scenario_action.triggered.connect(self.load_scenario)

        self.reload_scenario_action = QAction("Reload scenario...", self)
        self.reload_scenario_action.setIcon(self.style().standardIcon(QStyle.SP_BrowserReload))
        self.scenario_menu.addAction(self.reload_scenario_action)
        self.reload_scenario_action.triggered.connect(self.reload_scenario)

        self.save_scenario_action = QAction("Save scenario", self)
        self.save_scenario_action.setIcon(QIcon(":/icons/save_file.png"))
        self.scenario_menu.addAction(self.save_scenario_action)
        self.save_scenario_action.triggered.connect(self.save_scenario)

        save_scenario_as_action = QAction("Save scenario as...", self)
        save_scenario_as_action.setIcon(QIcon(":/icons/save_file_as.png"))
        self.scenario_menu.addAction(save_scenario_as_action)
        save_scenario_as_action.triggered.connect(self.save_scenario_as)

        undo_scenario_action = QAction("Undo scenario flux values edit", self)
        undo_scenario_action.setIcon(QIcon(":/icons/undo.png"))
        self.scenario_menu.addAction(undo_scenario_action)
        undo_scenario_action.triggered.connect(self.undo_scenario_edit)

        redo_scenario_action = QAction("Redo scenario flux values edit", self)
        redo_scenario_action.setIcon(QIcon(":/icons/redo.png"))
        self.scenario_menu.addAction(redo_scenario_action)
        redo_scenario_action.triggered.connect(self.redo_scenario_edit)

        clear_scenario_action = QAction("Clear scenario", self)
        self.scenario_menu.addAction(clear_scenario_action)
        clear_scenario_action.triggered.connect(self.clear_scenario)

        clear_all_action = QAction("Clear all reaction box entries", self)
        clear_all_action.setIcon(QIcon(":/icons/clear.png"))
        self.scenario_menu.addAction(clear_all_action)
        clear_all_action.triggered.connect(self.clear_all)

        add_values_to_scenario_action = QAction(
            "Add all flux values to scenario", self)
        self.scenario_menu.addAction(add_values_to_scenario_action)
        add_values_to_scenario_action.triggered.connect(
            self.add_values_to_scenario)

        merge_scenario_action = QAction("Merge flux values from scenario...", self)
        self.scenario_menu.addAction(merge_scenario_action)
        merge_scenario_action.triggered.connect(self.merge_scenario)

        set_model_bounds_to_scenario_action = QAction(
            "Use scenario flux values as model bounds", self)
        self.scenario_menu.addAction(set_model_bounds_to_scenario_action)
        set_model_bounds_to_scenario_action.triggered.connect(
            self.set_model_bounds_to_scenario)

        pin_scenario_reactions_action = QAction(
            "Pin all scenario reactions to top of reaction list", self)
        self.scenario_menu.addAction(pin_scenario_reactions_action)
        pin_scenario_reactions_action.triggered.connect(
            self.pin_scenario_reactions)

        unpin_all_reactions_action = QAction(
            "Unpin all reactions in reaction list", self)
        self.scenario_menu.addAction(unpin_all_reactions_action)
        unpin_all_reactions_action.triggered.connect(
            self.centralWidget().reaction_list.unpin_all)

        self.scenario_menu.addSeparator()

        update_action = QAction("Default Coloring", self)
        update_action.setIcon(QIcon(":/icons/default-color.png"))
        update_action.triggered.connect(self.central_widget.update)

        self.scenario_menu.addAction(self.heaton_action)
        self.scenario_menu.addAction(self.onoff_action)
        self.scenario_menu.addAction(update_action)

        self.clipboard_menu = self.menu.addMenu("Clipboard")

        copy_to_clipboard_action = QAction("Copy to clipboard", self)
        self.clipboard_menu.addAction(copy_to_clipboard_action)
        copy_to_clipboard_action.triggered.connect(self.copy_to_clipboard)

        paste_clipboard_action = QAction("Paste clipboard", self)
        self.clipboard_menu.addAction(paste_clipboard_action)
        paste_clipboard_action.triggered.connect(self.paste_clipboard)

        clipboard_arithmetics_action = QAction(
            "Clipboard arithmetics...", self)
        self.clipboard_menu.addAction(clipboard_arithmetics_action)
        clipboard_arithmetics_action.triggered.connect(
            self.clipboard_arithmetics)

        save_fluxes_as_csv_action = QAction("Save flux solution as comma-separated .csv...", self)
        self.clipboard_menu.addAction(save_fluxes_as_csv_action)
        save_fluxes_as_csv_action.triggered.connect(self.save_fluxes_as_csv)

        save_fluxes_as_xlsx_action = QAction("Save flux solution as Excel .xlsx...", self)
        self.clipboard_menu.addAction(save_fluxes_as_xlsx_action)
        save_fluxes_as_xlsx_action.triggered.connect(self.save_fluxes_as_xlsx)

        self.map_menu = self.menu.addMenu("Map")
        self.cnapy_map_actions = QActionGroup(self)
        separator = QAction(" CNApy map", self)
        separator.setSeparator(True)
        self.cnapy_map_actions.addAction(separator)
        self.escher_map_actions = QActionGroup(self)
        separator = QAction(" Escher map", self)
        separator.setSeparator(True)
        self.escher_map_actions.addAction(separator)

        add_map_action = QAction("Add new map", self)
        self.map_menu.addAction(add_map_action)
        add_map_action.triggered.connect(self.central_widget.add_map)

        add_escher_map_action = QAction("Add new map from Escher SVG...", self)
        self.map_menu.addAction(add_escher_map_action)
        add_escher_map_action.triggered.connect(self.add_escher_map)

        open_escher = QAction("Add interactive Escher map", self)
        self.map_menu.addAction(open_escher)
        open_escher.triggered.connect(lambda: self.central_widget.add_map(escher=True))

        self.change_map_name_action = QAction("Change map name", self)
        self.map_menu.addAction(self.change_map_name_action)
        self.change_map_name_action.triggered.connect(self.change_map_name)
        self.change_map_name_action.setEnabled(False)

        self.change_background_action = QAction("Change map background", self)
        self.change_background_action.triggered.connect(self.change_background)
        self.change_background_action.setEnabled(False)
        self.cnapy_map_actions.addAction(self.change_background_action)

        self.inc_bg_size_action = QAction("Increase background size", self)
        self.inc_bg_size_action.setShortcut("Ctrl+Shift++")
        self.inc_bg_size_action.triggered.connect(self.inc_bg_size)
        self.inc_bg_size_action.setEnabled(False)
        self.cnapy_map_actions.addAction(self.inc_bg_size_action)

        self.dec_bg_size_action = QAction("Decrease background size", self)
        self.dec_bg_size_action.setShortcut("Ctrl+Shift+-")
        self.dec_bg_size_action.triggered.connect(self.dec_bg_size)
        self.dec_bg_size_action.setEnabled(False)
        self.cnapy_map_actions.addAction(self.dec_bg_size_action)

        load_maps_action = QAction("Load reaction box positions...", self)
        load_maps_action.triggered.connect(self.load_box_positions)
        self.cnapy_map_actions.addAction(load_maps_action)

        self.save_box_positions_action = QAction(
            "Save reaction box positions...", self)
        self.save_box_positions_action.triggered.connect(
            self.save_box_positions)
        self.save_box_positions_action.setEnabled(False)
        self.cnapy_map_actions.addAction(self.save_box_positions_action)

        self.inc_box_size_action = QAction("Increase box size", self)
        self.inc_box_size_action.setShortcut("Ctrl++")
        self.inc_box_size_action.triggered.connect(self.inc_box_size)
        self.inc_box_size_action.setEnabled(False)
        self.cnapy_map_actions.addAction(self.inc_box_size_action)

        self.dec_box_size_action = QAction("Decrease box size", self)
        self.dec_box_size_action.setShortcut("Ctrl+-")
        self.dec_box_size_action.triggered.connect(self.dec_box_size)
        self.dec_box_size_action.setEnabled(False)
        self.cnapy_map_actions.addAction(self.dec_box_size_action)

        self.cnapy_screenshot_action = QAction("Take map view screenshot...", self)
        self.cnapy_screenshot_action.triggered.connect(self.take_screenshot_cnapy)
        self.cnapy_map_actions.addAction(self.cnapy_screenshot_action)

        escher_export_svg_action = QAction("Export as SVG...")
        escher_export_svg_action.triggered.connect(
            lambda: self.centralWidget().map_tabs.currentWidget().page().runJavaScript("builder.map.save_svg()"))
        self.escher_map_actions.addAction(escher_export_svg_action)

        escher_export_png_action = QAction("Export as PNG...")
        escher_export_png_action.triggered.connect(
            lambda: self.centralWidget().map_tabs.currentWidget().page().runJavaScript("builder.map.save_png()"))
        self.escher_map_actions.addAction(escher_export_png_action)

        escher_zoom_canvas_action = QAction("Zoom to canvas")
        escher_zoom_canvas_action.triggered.connect(
            lambda: self.centralWidget().map_tabs.currentWidget().page().runJavaScript("builder.map.zoom_extent_canvas()"))
        self.escher_map_actions.addAction(escher_zoom_canvas_action)

        # does not work as expected (TODO: why?), for now save JSON via Escher menu in edit mode
        # escher_save_map_action = QAction("Save map JSON...")
        # escher_save_map_action.triggered.connect(
        #     lambda: self.centralWidget().map_tabs.currentWidget().page().runJavaScript("builder.map.saveMap()"))
        # self.escher_map_actions.addAction(escher_save_map_action)

        escher_settings_action = QAction("Escher settings...")
        escher_settings_action.triggered.connect(
            lambda: self.centralWidget().map_tabs.currentWidget().page().runJavaScript(r"builder.passPropsSettingsMenu({display: true})"))
        self.escher_map_actions.addAction(escher_settings_action)

        self.escher_edit_mode_action = QAction("Edit mode")
        self.escher_edit_mode_action.triggered.connect(self.set_escher_edit_mode)
        self.escher_edit_mode_action.setCheckable(True)
        self.escher_map_actions.addAction(self.escher_edit_mode_action)

        self.map_menu.addActions(self.cnapy_map_actions.actions())
        self.escher_map_actions.setVisible(False)
        self.map_menu.addActions(self.escher_map_actions.actions())


        self.analysis_menu = self.menu.addMenu("Analysis")

        fba_action = QAction("Flux Balance Analysis (FBA)", self)
        fba_action.triggered.connect(self.fba)
        self.analysis_menu.addAction(fba_action)

        self.auto_fba_action = QAction("Auto FBA", self)
        self.auto_fba_action.triggered.connect(self.auto_fba)
        self.auto_fba_action.setCheckable(True)
        self.analysis_menu.addAction(self.auto_fba_action)

        pfba_action = QAction(
            "Parsimonious Flux Balance Analysis (pFBA)", self)
        pfba_action.triggered.connect(self.pfba)
        self.analysis_menu.addAction(pfba_action)

        fva_action = QAction("Flux Variability Analysis (FVA)", self)
        fva_action.triggered.connect(self.fva)
        self.analysis_menu.addAction(fva_action)

        make_scenario_feasible_action = QAction("Make scenario feasible...", self)
        make_scenario_feasible_action.triggered.connect(self.make_scenario_feasible)
        self.analysis_menu.addAction(make_scenario_feasible_action)
        self.make_scenario_feasible_dialog = None

        self.analysis_menu.addSeparator()

        self.efm_menu = self.analysis_menu.addMenu("Elementary Flux Modes")

        self.efmtool_action = QAction(
            "Compute Elementary Flux Modes via EFMtool...", self)
        self.efmtool_action.triggered.connect(self.efmtool)
        self.efm_menu.addAction(self.efmtool_action)

        load_modes_action = QAction("Load modes...", self)
        self.efm_menu.addAction(load_modes_action)
        load_modes_action.triggered.connect(self.load_modes)

        self.sd_menu = self.analysis_menu.addMenu("Computational Strain Design")
        sd_action = QAction("Compute Strain Designs...", self)
        sd_action.triggered.connect(self.strain_design)
        self.sd_menu.addAction(sd_action)
        self.sd_dialog = None
        self.sd_sols = None

        load_sd_action = QAction("Load Strain Designs...", self)
        self.sd_menu.addAction(load_sd_action)
        load_sd_action.triggered.connect(self.load_strain_designs)

        sd_action = QAction("Compute Minimal Cut Sets (legacy)...", self)
        sd_action.triggered.connect(self.mcs)
        self.sd_menu.addAction(sd_action)
        self.mcs_dialog = None

        load_mcs_action = QAction("Load Minimal Cut Sets (legacy)...", self)
        self.sd_menu.addAction(load_mcs_action)
        load_mcs_action.triggered.connect(self.load_mcs)

        self.flux_optimization_action = QAction(
            "Flux optimization...", self)
        self.flux_optimization_action.triggered.connect(self.optimize_flux)
        self.analysis_menu.addAction(self.flux_optimization_action)

        self.yield_optimization_action = QAction(
            "Yield optimization...", self)
        self.yield_optimization_action.triggered.connect(self.optimize_yield)
        self.analysis_menu.addAction(self.yield_optimization_action)

        plot_space_action = QAction("Plot phase plane/yield space...", self)
        plot_space_action.triggered.connect(self.plot_space)
        self.analysis_menu.addAction(plot_space_action)

        self.thermodynamic_menu = self.analysis_menu.addMenu("Thermodynamic analyses")

        optmdf_action = QAction("OptMDFpathway...", self)
        optmdf_action.triggered.connect(self.perform_optmdfpathway)
        self.thermodynamic_menu.addAction(optmdf_action)

        tfba_action = QAction("Thermodynamic FBA...", self)
        tfba_action.triggered.connect(self.perform_thermodynamic_fba)
        self.thermodynamic_menu.addAction(tfba_action)

        bottleneck_action = QAction("Thermodynamic bottleneck analysis...", self)
        bottleneck_action.triggered.connect(self.perform_bottleneck_analysis)
        self.thermodynamic_menu.addAction(bottleneck_action)

        self.thermodynamic_menu.addSeparator()

        dG0_menu = self.thermodynamic_menu.addMenu("Load dG'° values [in kJ/mol] (replacing all current values)...")

        dG0_json_action = QAction("...as JSON...", self)
        dG0_json_action.triggered.connect(self.load_dG0_json_replace_all)
        dG0_menu.addAction(dG0_json_action)

        dG0_xlsx_action = QAction("...as Excel XLSX...", self)
        dG0_xlsx_action.triggered.connect(self.load_dG0_xlsx_replace_all)
        dG0_menu.addAction(dG0_xlsx_action)

        dG0_menu = self.thermodynamic_menu.addMenu("Load dG'° values [in kJ/mol] (amending current values)...")

        dG0_json_action = QAction("...as JSON...", self)
        dG0_json_action.triggered.connect(self.load_dG0_json_amend)
        dG0_menu.addAction(dG0_json_action)

        dG0_xlsx_action = QAction("...as Excel XLSX...", self)
        dG0_xlsx_action.triggered.connect(self.load_dG0_xlsx_amend)
        dG0_menu.addAction(dG0_xlsx_action)

        concentrations_menu = self.thermodynamic_menu.addMenu("Load concentration ranges [in M] (replacing all current values)...")

        concentrations_json_action = QAction("...as JSON...", self)
        concentrations_json_action.triggered.connect(self.load_concentrations_json_replace_all)
        concentrations_menu.addAction(concentrations_json_action)

        concentrations_xlsx_action = QAction("...as Excel XLSX...", self)
        concentrations_xlsx_action.triggered.connect(self.load_concentrations_xlsx_replace_all)
        concentrations_menu.addAction(concentrations_xlsx_action)


        concentrations_menu = self.thermodynamic_menu.addMenu("Load concentration ranges [in M] (amending current values)...")

        concentrations_json_action = QAction("...as JSON...", self)
        concentrations_json_action.triggered.connect(self.load_concentrations_json_amend)
        concentrations_menu.addAction(concentrations_json_action)

        concentrations_xlsx_action = QAction("...as Excel XLSX...", self)
        concentrations_xlsx_action.triggered.connect(self.load_concentrations_xlsx_amend)
        concentrations_menu.addAction(concentrations_xlsx_action)

        self.analysis_menu.addSeparator()

        show_model_stats_action = QAction("Show model stats", self)
        self.analysis_menu.addAction(show_model_stats_action)
        show_model_stats_action.triggered.connect(
            self.execute_print_model_stats)

        show_model_bounds_action = QAction("Show flux bounds in reaction boxes", self)
        self.analysis_menu.addAction(show_model_bounds_action)
        show_model_bounds_action.triggered.connect(self.show_model_bounds)

        net_conversion_action = QAction(
            "Net conversion of external metabolites", self)
        self.analysis_menu.addAction(net_conversion_action)
        net_conversion_action.triggered.connect(
            self.show_net_conversion)

        all_in_out_fluxes_action = QAction(
            "Export all in/out fluxes as an XLSX table...", self)
        all_in_out_fluxes_action.triggered.connect(self.all_in_out_fluxes)
        self.analysis_menu.addAction(all_in_out_fluxes_action)

        in_out_flux_action = QAction(
            "Compute in/out fluxes at single metabolite...", self)
        in_out_flux_action.triggered.connect(self.in_out_flux)
        self.analysis_menu.addAction(in_out_flux_action)


        self.config_menu = self.menu.addMenu("Config")

        config_action = QAction("Configure CNApy...", self)
        config_action.setMenuRole(QAction.NoRole)
        self.config_menu.addAction(config_action)
        config_action.triggered.connect(self.show_config_dialog)

        config_action = QAction("Configure COBRApy...", self)
        config_action.setMenuRole(QAction.NoRole)
        self.config_menu.addAction(config_action)
        config_action.triggered.connect(self.show_config_cobrapy_dialog)

        config_action = QAction("Configure IBM CPLEX Full Version...", self)
        config_action.setMenuRole(QAction.NoRole)
        self.config_menu.addAction(config_action)
        config_action.triggered.connect(self.show_cplex_configuration_dialog)

        config_action = QAction("Configure Gurobi Full Version...", self)
        config_action.setMenuRole(QAction.NoRole)
        self.config_menu.addAction(config_action)
        config_action.triggered.connect(self.show_gurobi_configuration_dialog)

        show_console_action = QAction("Show Console", self)
        self.config_menu.addAction(show_console_action)
        show_console_action.triggered.connect(self.show_console)

        show_map_view = QAction("Show map view", self)
        self.config_menu.addAction(show_map_view)
        show_map_view.triggered.connect(self.show_map_view)

        show_model_view_action = QAction("Show model view", self)
        self.config_menu.addAction(show_model_view_action)
        show_model_view_action.triggered.connect(self.show_model_view)

        about_action = QAction("About CNApy...", self)
        about_action.setMenuRole(QAction.NoRole)
        self.config_menu.addAction(about_action)
        about_action.triggered.connect(self.show_about)

        zoom_in_action = QAction("Zoom in Map", self)
        zoom_in_action.setIcon(QIcon(":/icons/zoom-in.png"))
        zoom_in_action.triggered.connect(self.zoom_in)

        zoom_out_action = QAction("Zoom out Map", self)
        zoom_out_action.setIcon(QIcon(":/icons/zoom-out.png"))
        zoom_out_action.triggered.connect(self.zoom_out)

        self.set_current_filename("Untitled project")

        self.heaton_action.setCheckable(True)
        self.onoff_action.setCheckable(True)
        update_action.setCheckable(True)
        update_action.setChecked(True)
        self.colorings = QActionGroup(self)
        self.colorings.addAction(self.heaton_action)
        self.colorings.addAction(self.onoff_action)
        self.colorings.addAction(update_action)
        self.colorings.setExclusive(True)

        self.tool_bar = QToolBar()

        # load scenario action for tool bar
        self.load_scenario_action_tb = QAction("Load scenario...", self)
        self.load_scenario_action_tb.setToolTip("Load scenario file")
        self.load_scenario_action_tb.triggered.connect(self.load_scenario)
        self.tool_bar.addAction(self.load_scenario_action_tb)

        self.tool_bar.addAction(self.reload_scenario_action)
        self.tool_bar.addAction(self.save_scenario_action)
        self.tool_bar.addAction(save_scenario_as_action)
        self.tool_bar.addSeparator()
        self.tool_bar.addAction(clear_all_action)
        self.tool_bar.addAction(undo_scenario_action)
        self.tool_bar.addAction(redo_scenario_action)
        self.tool_bar.addSeparator()
        self.tool_bar.addAction(self.heaton_action)
        self.tool_bar.addAction(self.onoff_action)
        self.tool_bar.addAction(update_action)
        self.tool_bar.addSeparator()
        self.tool_bar.addAction(zoom_in_action)
        self.tool_bar.addAction(zoom_out_action)
        self.addToolBar(self.tool_bar)

        self.focus_search_action = QShortcut(
            QKeySequence('Ctrl+f'), self)
        self.focus_search_action.activated.connect(self.focus_search_box)

        status_bar: QStatusBar = self.statusBar()
        self.solver_status_display = QLabel()
        status_bar.addPermanentWidget(self.solver_status_display)
        self.solver_status_symbol = QLabel()
        status_bar.addPermanentWidget(self.solver_status_symbol)

        self.update_scenario_file_name()
        self.centralWidget().map_tabs.currentChanged.connect(self.on_tab_change)

    def closeEvent(self, event):
        if self.checked_unsaved():
            self.close_project_dialogs()
            # make sure Escher pages are destroyed before their profile
            self.delete_maps()
            event.accept()
            # releases the memory map file if this is a FluxVectorMemmap
            self.appdata.project.modes.clear()
        else:
            event.ignore()

    def checked_unsaved(self) -> bool:
        # TODO: check for changes in Escher maps instead of just setting unsaved changes
        # when acticvating the edit mode on an Escher map
        if self.appdata.unsaved:
            msgBox = QMessageBox()
            msgBox.setText("The project has been modified.")
            msgBox.setInformativeText("Do you want to save your changes?")
            msgBox.setStandardButtons(
                QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel)
            msgBox.setDefaultButton(QMessageBox.Save)
            ret = msgBox.exec_()
            if ret == QMessageBox.Save:
                # Save was clicked
                self.save_project_as()
                return True
            if ret == QMessageBox.Discard:
                # Don't save was clicked
                return True
            if ret == QMessageBox.Cancel:
                return False
        return True

    def unsaved_changes(self):
        if not self.appdata.unsaved:
            self.appdata.unsaved = True
            self.save_project_action.setEnabled(True)
            if len(self.appdata.project.name) == 0:
                shown_name = "Untitled project"
            else:
                shown_name = QFileInfo(self.appdata.project.name).fileName()

            self.setWindowTitle("CNApy - " + shown_name + ' - unsaved changes')

    def nounsaved_changes(self):
        if self.appdata.unsaved:
            self.appdata.unsaved = False
            self.save_project_action.setEnabled(False)
            if len(self.appdata.project.name) == 0:
                shown_name = "Untitled project"
            else:
                shown_name = QFileInfo(self.appdata.project.name).fileName()

            self.setWindowTitle("CNApy - " + shown_name)

    def disable_enable_dependent_actions(self):
        self.efm_action.setEnabled(False)

    @Slot()
    def exit_app(self):
        if self.checked_unsaved():
            # releases the memory map file if this is a FluxVectorMemmap
            self.appdata.project.modes.clear()
            QApplication.quit()

    def set_current_filename(self, filename):
        self.appdata.project.name = filename

        if len(self.appdata.project.name) == 0:
            shown_name = "Untitled project"
        else:
            shown_name = QFileInfo(self.appdata.project.name).fileName()

        self.setWindowTitle("CNApy - " + shown_name)

    @Slot()
    def show_about(self):
        dialog = AboutDialog(self.appdata)
        dialog.exec_()

    @Slot()
    def plot_space(self):
        self.plot_space = PlotSpaceDialog(self.appdata)
        self.plot_space.show()

    # Strain design computation and viewing functions
    def strain_design(self):
        self.sd_dialog = SDDialog(self.appdata)
        self.sd_dialog.show()

    @Slot(str)
    def strain_design_with_setup(self, sd_setup):
        self.sd_dialog = SDDialog(self.appdata, json.loads(sd_setup))
        self.sd_dialog.show()

    @Slot(str)
    def compute_strain_design(self,sd_setup):
        # launch progress viewer and computation thread
        self.sd_viewer = SDComputationViewer(self, self.appdata, sd_setup)
        self.sd_viewer.show_sd_signal.connect(self.show_strain_designs_with_setup, Qt.QueuedConnection)
        # connect signals to update progress
        self.sd_computation = SDComputationThread(self.appdata, sd_setup)
        self.sd_computation.output_connector.connect(self.sd_viewer.receive_progress_text, Qt.QueuedConnection)
        self.sd_computation.finished_computation.connect(self.sd_viewer.conclude_computation, Qt.QueuedConnection)
        self.sd_viewer.cancel_computation.connect(self.terminate_strain_design_computation)
        # show dialog and launch process
        # self.sd_viewer.exec()
        self.sd_viewer.show()
        self.sd_computation.start()

    def open_selected_recent_project(self):
        selected_last_project = self.sender().text()
        if not os.path.exists(selected_last_project):
            QMessageBox.critical(
                self,
                "File not found",
                "The selected recently opened .cna file could not be found. Possible reasons: The file was deleted, moved or renamed."
            )
            return
        if selected_last_project.endswith(".cna"):
            if self.checked_unsaved():
                self.open_project(filename=selected_last_project)
        else:
            self.new_project_from_sbml(filename=selected_last_project) # calls self.checked_unsaved()

    def update_recently_used_models(self, filename: str):
        if filename in self.appdata.recent_cna_files:
            filename_index = self.appdata.recent_cna_files.index(filename)
            del(self.appdata.recent_cna_files[filename_index])
        if len(self.appdata.recent_cna_files) > 19:  # Actually allows 20 shown recent .cna files
            del(self.appdata.recent_cna_files[-1])
        self.appdata.recent_cna_files.insert(0, filename)
        self.appdata.save_cnapy_config()
        self.build_recent_cna_menu()

    def build_recent_cna_menu(self):
        recent_cnas = list(self.recent_cna_actions.keys())
        for recent_cna in recent_cnas:
            self.recent_cna_menu.removeAction(self.recent_cna_actions[recent_cna])
            del(self.recent_cna_actions[recent_cna])

        for recent_cna in self.appdata.recent_cna_files:
            self.recent_cna_actions[recent_cna] = QAction(recent_cna, self)
            self.recent_cna_actions[recent_cna].triggered.connect(self.open_selected_recent_project)
            self.recent_cna_menu.addAction(self.recent_cna_actions[recent_cna])

    @Slot()
    def terminate_strain_design_computation(self):
        self.sd_computation.output_connector.disconnect()
        self.sd_computation.finished_computation.disconnect()
        self.sd_computation.terminate()

    @Slot(bytes)
    def show_strain_designs_with_setup(self, solutions_with_setup):
        self.show_strain_designs(solutions_with_setup, with_setup=True)

    @Slot(bytes)
    def show_strain_designs(self, solutions, with_setup=False):
        self.sd_sols = SDViewer(self.appdata, solutions, with_setup)
        self.sd_sols.show()
        self.centralWidget().update_mode()

    @Slot()
    def load_strain_designs(self):
        dialog = QFileDialog(self)
        filename: str = dialog.getOpenFileName(
            directory=self.appdata.work_directory, filter="*.sds")[0]
        if not filename or len(filename) == 0 or not os.path.exists(filename):
            return
        with open(filename,'rb') as f:
            solutions = f.read()
        self.show_strain_designs(solutions)

    @Slot()
    def optimize_yield(self):
        dialog = YieldOptimizationDialog(self.appdata, self.centralWidget())
        dialog.exec_()

    @Slot()
    def optimize_flux(self):
        dialog = FluxOptimizationDialog(self.appdata, self.centralWidget())
        dialog.exec_()

    @Slot()
    def show_config_dialog(self, first_start=False):
        dialog = ConfigDialog(self, first_start)
        if not first_start:
            dialog.exec_()

    @Slot()
    def show_config_cobrapy_dialog(self):
        dialog = ConfigCobrapyDialog(self.appdata)
        if self.mcs_dialog is not None:
            dialog.optlang_solver_set.connect(self.mcs_dialog.set_optlang_solver_text)
            dialog.optlang_solver_set.connect(self.mcs_dialog.configure_solver_options)
        if self.sd_dialog is not None:
            dialog.optlang_solver_set.connect(self.sd_dialog.set_optlang_solver_text)
            dialog.optlang_solver_set.connect(self.sd_dialog.configure_solver_options)
        dialog.exec_()

    @Slot()
    def show_cplex_configuration_dialog(self):
        dialog = CplexConfigurationDialog(self.appdata)
        dialog.exec_()

    @Slot()
    def show_gurobi_configuration_dialog(self):
        dialog = GurobiConfigurationDialog(self.appdata)
        dialog.exec_()

    @Slot()
    def export_sbml(self):
        dialog = QFileDialog(self)
        filename: str = dialog.getSaveFileName(
            directory=self.appdata.work_directory, filter=SBML_suffixes)[0]
        if not filename or len(filename) == 0:
            return

        self.setCursor(Qt.BusyCursor)
        try:
            self.save_sbml(filename)
        except ValueError:
            exstr = get_last_exception_string()
            utils.show_unknown_error_box(exstr)

        self.setCursor(Qt.ArrowCursor)

    @Slot()
    def download_examples(self):
        dialog = DownloadDialog(self.appdata)
        dialog.exec_()

    @Slot()
    def load_box_positions(self):
        dialog = QFileDialog(self)
        filename: str = dialog.getOpenFileName(
            directory=self.appdata.work_directory, filter="*.maps")[0]
        if not filename or len(filename) == 0 or not os.path.exists(filename):
            return

        idx = self.centralWidget().map_tabs.currentIndex()
        if idx < 0:
            self.centralWidget().add_map()
            idx = self.centralWidget().map_tabs.currentIndex()
        name = self.centralWidget().map_tabs.tabText(idx)

        with open(filename, 'r') as fp:
            self.appdata.project.maps[name]["boxes"] = json.load(fp)

        to_remove = []
        for r_id in self.appdata.project.maps[name]["boxes"].keys():
            if not self.appdata.project.cobra_py_model.reactions.has_id(r_id):
                to_remove.append(r_id)

        for r_id in to_remove:
            self.appdata.project.maps[name]["boxes"].pop(r_id)

        self.recreate_maps()
        self.unsaved_changes()
        self.centralWidget().update()

    @Slot()
    def merge_scenario(self):
        self.load_scenario(merge=True)

    @Slot()
    def load_scenario(self, merge=False):
        dialog = QFileDialog(self)
        filename: str = dialog.getOpenFileName(
            caption="Load scenario",
            directory=self.appdata.last_scen_directory,
            filter="*.scen *.val"
        )[0]
        if not filename or len(filename) == 0 or not os.path.exists(filename):
            return
        self.load_scenario_file(filename, merge=merge)

    def load_scenario_file(self, filename, merge=False):
        self.appdata.scenario_past.clear()
        self.appdata.scenario_future.clear()
        try:
            missing_reactions, incompatible_constraints, skipped_scenario_reactions = \
                self.appdata.project.scen_values.load(filename, self.appdata, merge=merge)
        except json.decoder.JSONDecodeError:
            QMessageBox.critical(
                self,
                'Could not open file',
                "File could not be opened as it does not seem to be a valid scenario file. "
                "Maybe the file got the .scen ending for other reasons than being a scenario file or the file is corrupted."
            )
            return

        self.centralWidget().reaction_list.pin_multiple(self.appdata.project.scen_values.pinned_reactions)

        self.appdata.project.comp_values.clear()
        self.appdata.project.fva_values.clear()
        self.central_widget.tabs.widget(ModelTabIndex.Scenario).recreate_scenario_items()
        self.appdata.project.update_reaction_id_lists()

        if len(missing_reactions) > 0 :
            QMessageBox.warning(self, 'Unknown reactions in scenario',
            'The following reaction IDs of the scenario do not exist in the current model and will be ignored:\n'+' '.join(missing_reactions))

        if len(skipped_scenario_reactions) > 0 :
            QMessageBox.warning(self, 'Reactions with existing IDs in scenario',
            'The scenario reactions with the following IDs already exist in the current model and will be ignored:\n'+' '.join(skipped_scenario_reactions))

        if len(incompatible_constraints) > 0 :
            QMessageBox.warning(self, 'Unknown reactions in scenario',
            'The following scenario constraints refer to reactions not in the model and will be ignored:\n'+
            '\n'.join([utils.format_scenario_constraint(c) for c in incompatible_constraints]))

        if self.appdata.auto_fba:
            self.fba()
        else:
            self.centralWidget().update()
            self.clear_status_bar()
        self.appdata.last_scen_directory = os.path.dirname(filename)
        self.appdata.project.scen_values.has_unsaved_changes = False
        self.update_scenario_file_name()

    @Slot()
    def load_modes(self):
        dialog = QFileDialog(self)
        filename: str = dialog.getOpenFileName(
            directory=self.appdata.work_directory, filter="*.npz")[0]
        if not filename or len(filename) == 0 or not os.path.exists(filename):
            return

        self.appdata.project.modes = FluxVectorContainer(filename)
        self.centralWidget().mode_navigator.current = 0

        self.centralWidget().mode_navigator.set_to_efm()
        self.centralWidget().update_mode()

    @Slot()
    def load_mcs(self):
        dialog = QFileDialog(self)
        filename: str = dialog.getOpenFileName(
            directory=self.appdata.work_directory, filter="*.npz")[0]
        if not filename or len(filename) == 0 or not os.path.exists(filename):
            return

        self.appdata.project.modes = FluxVectorContainer(filename)
        self.centralWidget().mode_navigator.current = 0
        self.centralWidget().mode_navigator.set_to_mcs()
        self.centralWidget().update_mode()

    @Slot()
    def change_background(self, caption="Select a SVG file", directory=None):
        '''Load a background image for the current map'''
        dialog = QFileDialog(self)
        filename: str = dialog.getOpenFileName(caption=caption,
            directory=self.appdata.work_directory if directory is None else directory,
            filter="*.svg")[0]
        if not filename or len(filename) == 0 or not os.path.exists(filename):
            return None

        idx = self.centralWidget().map_tabs.currentIndex()
        name = self.centralWidget().map_tabs.tabText(idx)

        self.appdata.project.maps[name]["background"] = filename
        # rebuild everything to make sure that background is displayed behind the boxes
        self.centralWidget().map_tabs.widget(idx).rebuild_scene()
        self.centralWidget().update()
        self.centralWidget().map_tabs.setCurrentIndex(idx)
        self.centralWidget().map_tabs.widget(idx).fit()
        self.unsaved_changes()

        return filename

    # the variant below requires both the Escher JSON and an exported SVG
    # @Slot()
    # def add_escher_map(self):
    #     dialog = QFileDialog(self)
    #     filename: str = dialog.getOpenFileName(caption="Select an Escher JSON file",
    #         directory=self.appdata.work_directory, filter="*.json")[0]
    #     if not filename or len(filename) == 0 or not os.path.exists(filename):
    #         return
    #
    #     with open(filename) as fp:
    #         map_json = json.load(fp)
    #
    #     map_name, map_idx = self.centralWidget().add_map(base_name=map_json[0]['map_name'])
    #     self.change_background(caption="Select a SVG file that corresponds to the previously loaded Escher JSON file")
    #
    #     reaction_bigg_ids = dict()
    #     for r in self.appdata.project.cobra_py_model.reactions:
    #         bigg_id = r.annotation.get("bigg.reaction", None)
    #         if bigg_id is None: # if there is no BiGG ID in the annotation...
    #             bigg_id = r.id # ... use the reaction ID as proxy
    #         reaction_bigg_ids[bigg_id] = r.id
    #
    #     offset_x = map_json[1]['canvas']['x']
    #     offset_y = map_json[1]['canvas']['y']
    #     self.appdata.project.maps[map_name]["boxes"] = dict()
    #     for r in map_json[1]["reactions"].values():
    #         bigg_id = r["bigg_id"]
    #         if bigg_id in reaction_bigg_ids:
    #             self.appdata.project.maps[map_name]["boxes"][reaction_bigg_ids[bigg_id]] = [r["label_x"] - offset_x, r["label_y"] - offset_y]
    #
    #     self.recreate_maps()
    #     self.centralWidget().map_tabs.setCurrentIndex(map_idx)

    @Slot()
    def add_escher_map(self, annotation_key_for_id="bigg.reaction", strip_compartment=False):
        # maps gets a default name because an Escher SVG file does not contain the map name
        has_unsaved_changes = self.appdata.unsaved
        map_name, map_idx = self.centralWidget().add_map()
        file_name = self.change_background(caption="Select an Escher SVG file")
        if file_name is None:
            del self.appdata.project.maps[map_name]
            self.centralWidget().map_tabs.removeTab(map_idx)
            self.appdata.unsaved = has_unsaved_changes
            return

        reaction_bigg_ids = dict()
        for r in self.appdata.project.cobra_py_model.reactions:
            bigg_id = r.annotation.get(annotation_key_for_id, None)
            if not isinstance(bigg_id, str): # if there is no (unique) BiGG ID in the annotation...
                bigg_id = r.id # ... use the reaction ID as proxy
            if strip_compartment:
                for c_id in self.appdata.project.cobra_py_model.compartments.keys():
                    if bigg_id.endswith(c_id):
                        bigg_id = bigg_id[:-(len(c_id)+1)] # +1 for the _
                        break
                # print(bigg_id, r.id)
            reaction_bigg_ids[bigg_id] = r.id

        def get_translate_coordinates(translate: str):
            x_y = translate.split("(")[1].split(",")
            return float(x_y[0]), float(x_y[1][:-1])
        self.appdata.project.maps[map_name]["boxes"] = dict()
        try:
            root = ET.parse(file_name).getroot()
            graph = root.find('{http://www.w3.org/2000/svg}g')
            canvas_group = None
            reactions = None
            for child in graph:
                # print(child.tag, child.attrib)
                if child.attrib.get("class", None) == "canvas-group":
                    canvas_group = child
                if child.attrib.get("id", None) == "reactions":
                    reactions = child
            if canvas_group is None or reactions is None:
                raise ValueError
            canvas = None
            for child in canvas_group:
                if child.attrib.get("id", None) == "canvas":
                    canvas = child
            if canvas is None:
                raise ValueError
            (offset_x, offset_y) = get_translate_coordinates(canvas.attrib['transform'])
            for r in reactions:
                for child in r:
                    if child.attrib.get("class", None) == "reaction-label-group":
                        for neighbor in child.iter():
                            if neighbor.attrib.get("class", None) == "reaction-label label":
                                bigg_id = neighbor.text
                                if bigg_id in reaction_bigg_ids:
                                    (label_x, label_y) =  get_translate_coordinates(child.attrib['transform'])
                                    self.appdata.project.maps[map_name]["boxes"][reaction_bigg_ids[bigg_id]] = [label_x - offset_x, label_y - offset_y]
        except:
            QMessageBox.critical(self, "Failed to parse "+file_name+" as Escher SVG file",
                                 file_name+" does not appear to have been exported from Escher. "
                                 "Automatic mapping of reaction boxes not possible.")
            self.appdata.project.maps[map_name]["boxes"] = dict()

        self.recreate_maps()
        self.centralWidget().map_tabs.setCurrentIndex(map_idx)
        self.centralWidget().map_tabs.widget(map_idx).fit()

    @Slot()
    def change_map_name(self):
        '''Execute RenameMapDialog'''
        dialog = RenameMapDialog(
            self.appdata, self.centralWidget())
        dialog.exec_()

    @Slot()
    def inc_box_size(self):
        idx = self.centralWidget().map_tabs.currentIndex()
        name = self.centralWidget().map_tabs.tabText(idx)
        self.appdata.project.maps[name]["box-size"] *= 1.1
        self.unsaved_changes()
        self.centralWidget().update()

    @Slot()
    def dec_box_size(self):
        idx = self.centralWidget().map_tabs.currentIndex()
        name = self.centralWidget().map_tabs.tabText(idx)
        self.appdata.project.maps[name]["box-size"] *= (1/1.1)
        self.unsaved_changes()
        self.centralWidget().update()

    @Slot()
    def inc_bg_size(self):
        idx = self.centralWidget().map_tabs.currentIndex()
        name = self.centralWidget().map_tabs.tabText(idx)
        self.appdata.project.maps[name]["bg-size"] *= 1.1
        self.unsaved_changes()
        self.centralWidget().update()

    @Slot()
    def dec_bg_size(self):
        idx = self.centralWidget().map_tabs.currentIndex()
        name = self.centralWidget().map_tabs.tabText(idx)
        self.appdata.project.maps[name]["bg-size"] *= (1/1.1)
        self.unsaved_changes()
        self.centralWidget().update()

    @Slot()
    def zoom_in(self):
        mv = self.centralWidget().map_tabs.currentWidget()
        if mv is not None:
            mv.zoom_in()

    @Slot()
    def zoom_out(self):
        mv = self.centralWidget().map_tabs.currentWidget()
        if mv is not None:
            mv.zoom_out()

    @Slot(bool)
    def set_escher_edit_mode(self, checked: bool):
        # cannot use the parameter checked because in ActionGroup the entries are mutually exclusive
        # and the policy cannot be changed in Python (why?)
        now_enabled = not self.centralWidget().map_tabs.currentWidget().editing_enabled
        self.centralWidget().map_tabs.currentWidget().enable_editing(now_enabled)
        self.escher_edit_mode_action.setChecked(now_enabled)
        self.unsaved_changes() # preliminary solution until checking for changes in Escher maps is implemented

    @Slot()
    def focus_search_box(self):
        self.centralWidget().searchbar.setFocus()

    @Slot()
    def save_box_positions(self):
        idx = self.centralWidget().map_tabs.currentIndex()
        name = self.centralWidget().map_tabs.tabText(idx)

        dialog = QFileDialog(self)
        filename: str = dialog.getSaveFileName(
            directory=self.appdata.work_directory, filter="*.maps")[0]
        if not filename or len(filename) == 0:
            return

        with open(filename, 'w') as fp:
            json.dump(self.appdata.project.maps[name]["boxes"], fp)

    @Slot()
    def save_scenario(self):
        self.appdata.project.scen_values.save(self.appdata.project.scen_values.file_name)
        self.update_scenario_file_name()

    @Slot()
    def save_scenario_as(self):
        filename: str = QFileDialog.getSaveFileName(
            directory=self.appdata.last_scen_directory, filter="*.scen")[0]
        if len(filename) > 0:
            if not filename.endswith(".scen"):
                filename += ".scen"
            self.appdata.project.scen_values.file_name = filename
            self.appdata.project.scen_values.save(filename)
            self.appdata.last_scen_directory = os.path.dirname(filename)
            self.update_scenario_file_name()

    def undo_scenario_edit(self):
        ''' undo last edit in scenario history '''
        if len(self.appdata.scenario_past) > 0:
            last = self.appdata.scenario_past.pop()
            self.appdata.scenario_future.append(last)
            self.appdata.recreate_scenario_from_history()
            if self.appdata.auto_fba:
                self.fba()
            self.centralWidget().update()

    def redo_scenario_edit(self):
        ''' redo last undo of scenario history '''
        if len(self.appdata.scenario_future) > 0:
            nex = self.appdata.scenario_future.pop()
            self.appdata.scenario_past.append(nex)
            self.appdata.recreate_scenario_from_history()
            if self.appdata.auto_fba:
                self.fba()
            self.centralWidget().update()

    def clear_scenario(self):
        self.appdata.scen_values_clear()
        self.appdata.project.scen_values.clear()
        self.update_scenario_file_name()
        self.central_widget.tabs.widget(ModelTabIndex.Scenario).recreate_scenario_items_needed = True
        if self.appdata.auto_fba:
            self.fba()
        self.centralWidget().update()

    def clear_all(self):
        self.appdata.scen_values_clear()
        self.update_scenario_file_name()
        self.appdata.project.comp_values.clear()
        self.appdata.project.comp_values_type = 0
        self.appdata.project.fva_values.clear()
        self.appdata.project.conc_values.clear()
        self.appdata.project.df_values.clear()
        self.appdata.project.high = 0
        self.appdata.project.low = 0
        self.centralWidget().update()
        self.clear_status_bar()

    def load_default_scenario(self):
        self.appdata.project.comp_values.clear()
        self.appdata.project.fva_values.clear()
        self.appdata.project.scen_values.clear()
        self.update_scenario_file_name()
        (reactions, values) = self.appdata.project.collect_default_scenario_values()
        if len(reactions) == 0:
            self.appdata.scen_values_clear()
        else:
            self.appdata.scen_values_set_multiple(reactions, values)
        if self.appdata.auto_fba:
            self.fba()
        else:
            self.centralWidget().update()
            self.clear_status_bar()

    @Slot()
    def new_project(self):
        if self.checked_unsaved():
            self.new_project_unchecked()
            self.recreate_maps()

    def new_project_unchecked(self):
        self.appdata.project = ProjectData()
        self.delete_maps()

        self.centralWidget().mode_navigator.clear()
        self.centralWidget().clear_model_item_history()
        self.centralWidget().reaction_list.reaction_list.clear()
        self.close_project_dialogs()

        self.appdata.project.scen_values.clear()
        self.appdata.scenario_past.clear()
        self.appdata.scenario_future.clear()

        self.set_current_filename("Untitled project")
        self.nounsaved_changes()

    @Slot()
    def new_project_from_sbml(self, filename=''):
        if self.checked_unsaved():
            if len(filename) == 0:
                dialog = QFileDialog(self)
                filename: str = dialog.getOpenFileName(
                    directory=self.appdata.work_directory, filter=SBML_suffixes)[0]
            if not filename or len(filename) == 0 or not os.path.exists(filename):
                return

            self.setCursor(Qt.BusyCursor)
            try:
                cobra_py_model = CNApyModel.read_sbml_model(filename)
            except cobra.io.sbml.CobraSBMLError:
                exstr = get_last_exception_string()
                QMessageBox.warning(
                    self, 'Could not read sbml.', exstr)
                return
            self.new_project_unchecked()
            self.appdata.project.cobra_py_model = cobra_py_model

            self.recreate_maps()
            self.centralWidget().update(rebuild_all_tabs=True)
            self.update_scenario_file_name()
            self.update_recently_used_models(filename)

            self.setCursor(Qt.ArrowCursor)

    def open_project(self, filename):
        self.close_project_dialogs()
        temp_dir = TemporaryDirectory()

        self.setCursor(Qt.BusyCursor)
        try:
            with ZipFile(filename, 'r') as zip_ref:
                zip_ref.extractall(temp_dir.name)

                box_positions_path = temp_dir.name+"/box_positions.json"
                if not os.path.exists(box_positions_path):
                    QMessageBox.critical(
                        self,
                        'Could not open file',
                        "File could not be opened as it does not seem to be a valid CNApy project, even though the file is a zip file. "
                        "Maybe the file got the .cna ending for other reasons than being a CNApy project or the file is corrupted."
                    )
                    self.setCursor(Qt.ArrowCursor)
                    return

                with open(box_positions_path, 'r') as fp:
                    maps = json.load(fp)

                    count = 1
                    for _name, m in maps.items():
                        m["background"] = temp_dir.name + \
                            "/map" + str(count) + ".svg"
                        count += 1
                # load meta_data
                with open(temp_dir.name+"/meta.json", 'r') as fp:
                    meta_data = json.load(fp)

                try:
                    cobra_py_model = CNApyModel.read_sbml_model(
                        temp_dir.name + "/model.sbml")
                except cobra.io.sbml.CobraSBMLError:
                    exstr = get_last_exception_string()
                    QMessageBox.warning(
                        self, 'Could not open project.', exstr)
                    return
                self.appdata.temp_dir = temp_dir
                self.appdata.project.maps = maps
                self.appdata.project.meta_data = meta_data
                self.appdata.project.cobra_py_model = cobra_py_model
                self.set_current_filename(filename)
                self.recreate_maps()
                self.centralWidget().mode_navigator.clear()
                self.centralWidget().clear_model_item_history()
                self.centralWidget().reaction_list.last_selected = None
                self.centralWidget().metabolite_list.last_selected = None
                self.centralWidget().gene_list.last_selected = None
                self.appdata.project.scen_values.clear()
                self.appdata.project.comp_values.clear()
                self.appdata.project.fva_values.clear()
                self.appdata.scenario_past.clear()
                self.appdata.scenario_future.clear()
                self.clear_status_bar()
                self.update_scenario_file_name()
                (reactions, values) = self.appdata.project.collect_default_scenario_values()
                if len(reactions) > 0:
                    self.appdata.scen_values_set_multiple(reactions, values)
                self.nounsaved_changes()

                # if project contains maps move splitter and fit mapview
                if len(self.appdata.project.maps) > 0:
                    (_, r) = self.centralWidget().splitter2.getRange(1)
                    self.centralWidget().splitter2.moveSplitter(round(r*0.8), 1)
                    self.centralWidget().fit_mapview()

                self.centralWidget().update(rebuild_all_tabs=True)
                self.update_scenario_file_name()
                self.update_recently_used_models(filename)

        except FileNotFoundError:
            exstr = get_last_exception_string()
            QMessageBox.warning(self, 'Could not open project.', exstr)
        except BadZipFile:
            QMessageBox.critical(
                self,
                'Could not open file',
                "File could not be opened as it does not seem to be a valid CNApy project. "
                "Maybe the file got the .cna ending for other reasons than being a CNApy project or the file is corrupted."
            )

        self.setCursor(Qt.ArrowCursor)

    @Slot()
    def open_project_dialog(self):
        if self.checked_unsaved():
            dialog = QFileDialog(self)
            filename: str = dialog.getOpenFileName(
                directory=self.appdata.work_directory, filter="*.cna")[0]
            if not filename or len(filename) == 0 or not os.path.exists(filename):
                return
            self.open_project(filename)

    def close_project_dialogs(self):
        '''closes modeless dialogs'''
        if self.mcs_dialog is not None:
            self.mcs_dialog.close()
            self.mcs_dialog = None
        if self.sd_dialog is not None:
            self.sd_dialog.close()
            self.sd_dialog = None
        if self.sd_sols is not None:
            self.sd_sols.close()
            self.sd_sols = None
        if self.make_scenario_feasible_dialog is not None:
            self.make_scenario_feasible_dialog.close()
            self.make_scenario_feasible_dialog = None

    def save_sbml(self, filename):
        '''Save model as SBML'''

        # cleanup to work around cobrapy not setting a default compartment
        # remove unused species - > cleanup disabled for now because of issues
        # with prune_unused_metabolites
        # (clean_model, unused_mets) = prune_unused_metabolites(
        #     self.appdata.project.cobra_py_model)
        clean_model = self.appdata.project.cobra_py_model
        # set unset compartments to ''
        undefined = ''
        for m in clean_model.metabolites:
            if m.compartment is None:
                undefined += m.id+'\n'
                m.compartment = 'undefined_compartment_please_fix'
            else:
                x = m.compartment
                x.strip()
                if x == '':
                    undefined += m.id+'\n'
                    m.compartment = 'undefined_compartment_please_fix'

        if undefined != '':
            QMessageBox.warning(self, 'Undefined compartments',
                                'The following metabolites have undefined compartments!\n' +
                                undefined+'\nPlease check!')

        self.appdata.project.cobra_py_model = clean_model

        cobra.io.write_sbml_model(
            self.appdata.project.cobra_py_model, filename)

    @Slot()
    def save_project(self):
        escher_map_count: int = 0
        semaphore = [0] # list with one integer to emulate pass by reference
        for i in range(len(self.centralWidget().map_tabs)):
            if isinstance(self.centralWidget().map_tabs.widget(i), EscherMapView):
                self.centralWidget().map_tabs.widget(i).retrieve_map_data(semaphore=semaphore)
                self.centralWidget().map_tabs.widget(i).retrieve_pos_and_zoom(semaphore=semaphore)
                escher_map_count += 1
        if escher_map_count > 0: # give some time for retrieve_map_data to finish
            escher_map_count *= 2
            timer = QTimer()
            wait_count = 0
            timer.setInterval(escher_map_count*10)
            def wait_for_retrieval():
                nonlocal wait_count
                if semaphore[0] == escher_map_count:
                    timer.stop()
                    self.continue_save_project()
                if wait_count >= 20:
                    timer.stop()
                    raise ValueError("Failed to retrieve Escher data, cannot save project.")
                wait_count += 1
            timer.timeout.connect(wait_for_retrieval)
            timer.start()
        else:
            self.continue_save_project()

    @Slot()
    def continue_save_project(self):
        ''' Save the project '''
        tmp_dir = TemporaryDirectory().name
        filename: str = self.appdata.project.name

        self.setCursor(Qt.BusyCursor)
        try:
            self.save_sbml(tmp_dir + "model.sbml")
        except ValueError:
            exstr = get_last_exception_string()
            utils.show_unknown_error_box(exstr)

            return

        svg_files = {}
        count = 1
        for name, m in self.appdata.project.maps.items():
            if m.get('view', 'cnapy') == 'cnapy':
                arc_name = "map" + str(count) + ".svg"
                svg_files[m["background"]] = arc_name
                m["background"] = arc_name
            count += 1

        # Save maps information
        # also contains the Escher map JSONs
        with open(tmp_dir + "box_positions.json", 'w') as fp:
            json.dump(self.appdata.project.maps, fp, skipkeys=True)

        # Save meta data
        self.appdata.project.meta_data["format version"] = self.appdata.format_version
        with open(tmp_dir + "meta.json", 'w') as fp:
            json.dump(self.appdata.project.meta_data, fp)

        with ZipFile(filename, 'w') as zip_obj:
            zip_obj.write(tmp_dir + "model.sbml", arcname="model.sbml")
            zip_obj.write(tmp_dir + "box_positions.json",
                          arcname="box_positions.json")
            zip_obj.write(tmp_dir + "meta.json", arcname="meta.json")
            for name, m in svg_files.items():
                zip_obj.write(name, arcname=m)

        # put svgs into temporary directory and update references
        with ZipFile(filename, 'r') as zip_ref:
            zip_ref.extractall(self.appdata.temp_dir.name)
            count = 1
            for name, m in self.appdata.project.maps.items():
                m["background"] = self.appdata.temp_dir.name + \
                    "/map" + str(count) + ".svg"
                count += 1

        self.nounsaved_changes()
        self.setCursor(Qt.ArrowCursor)

    @Slot()
    def save_project_as(self):
        filename: str = QFileDialog.getSaveFileName(
            directory=self.appdata.work_directory, filter="*.cna")[0]
        if len(filename) > 0:
            if not filename.endswith(".cna"):
                filename += ".cna"
            self.set_current_filename(filename)
            self.save_project()
            self.update_recently_used_models(filename)

    # TODO: are there really situations where _all_ maps need to be recreated?
    def recreate_maps(self):
        self.delete_maps()
        for name, mmap in self.appdata.project.maps.items():
            if mmap.get("view", "cnapy") == "cnapy":
                mmap = MapView(self.appdata, self.centralWidget(), name)
                mmap.show()
                self.centralWidget().connect_map_view_signals(mmap)
            elif mmap["view"] == "escher":
                mmap = EscherMapView(self.centralWidget(), name)
                self.centralWidget().connect_escher_map_view_signals(mmap)
                self.appdata.project.maps[name][EscherMapView] = mmap
            else:
                raise ValueError("Unknown map type "+mmap["view"])
            self.centralWidget().map_tabs.addTab(mmap, name)
            mmap.update()

    def delete_maps(self):
        with QSignalBlocker(self.centralWidget().map_tabs):
            for i in range(0, self.centralWidget().map_tabs.count()):
                self.centralWidget().map_tabs.widget(i).deleteLater()
            self.centralWidget().map_tabs.clear()

    def take_screenshot_cnapy(self):
        dialog = QFileDialog(self)
        filename: str = dialog.getSaveFileName(
            directory=self.appdata.work_directory,
            filter="*.png",
        )[0]
        if not filename or len(filename) == 0:
            return
        elif len(filename) <= 4 or filename[-4:] != ".png":
            filename += ".png"

        self.setCursor(Qt.BusyCursor)
        scale_factor = 10.0
        view = self.centralWidget().map_tabs.currentWidget()
        original_size = QSize(view.size())

        view.resize(original_size * scale_factor)
        view.setTransform(view.transform().scale(scale_factor, scale_factor))

        pixmap = view.grab()
        pixmap.save(filename, "PNG")

        view.setTransform(view.transform().scale(1/scale_factor, 1/scale_factor))
        view.resize(original_size)
        self.setCursor(Qt.ArrowCursor)

    def on_tab_change(self, idx):
        if idx >= 0:
            self.change_map_name_action.setEnabled(True)
            self.change_background_action.setEnabled(True)
            self.inc_box_size_action.setEnabled(True)
            self.dec_box_size_action.setEnabled(True)
            self.inc_bg_size_action.setEnabled(True)
            self.dec_bg_size_action.setEnabled(True)
            self.save_box_positions_action.setEnabled(True)
            self.centralWidget().update_map(idx)
            if isinstance(self.centralWidget().map_tabs.widget(idx), MapView):
                self.escher_map_actions.setVisible(False)
                self.cnapy_map_actions.setVisible(True)
                self.colorings.setEnabled(True)
                self.central_widget.search_annotations.setEnabled(True)
            else: # EscherMapView
                self.cnapy_map_actions.setVisible(False)
                self.escher_map_actions.setVisible(True)
                self.colorings.setEnabled(False)
                self.central_widget.search_annotations.setEnabled(False)
        else:
            self.change_map_name_action.setEnabled(False)
            self.change_background_action.setEnabled(False)
            self.inc_box_size_action.setEnabled(False)
            self.dec_box_size_action.setEnabled(False)
            self.inc_bg_size_action.setEnabled(False)
            self.dec_bg_size_action.setEnabled(False)
            self.save_box_positions_action.setEnabled(False)

    def copy_to_clipboard(self):
        self.appdata.clipboard_comp_values = self.appdata.project.comp_values.copy()

        for (key, value) in self.appdata.project.scen_values.items():
            self.appdata.clipboard_comp_values[key] = value

    def paste_clipboard(self):
        try:
            self.appdata.project.comp_values = self.appdata.clipboard_comp_values.copy()

            for key in (self.appdata.project.scen_values.keys() & self.appdata.clipboard_comp_values.keys()):
                self.appdata.project.scen_values[key] = self.appdata.clipboard_comp_values[key]
        except AttributeError:
            QMessageBox.warning(
                self,
                "No clipboard created yet",
                "Paste clipboard does not work as no clipboard was created yet. Store values to a clipboard first to solve this problem."
            )
            return
        self.centralWidget().update()

    @Slot()
    def clipboard_arithmetics(self):
        dialog = ClipboardCalculator(self.appdata)
        dialog.exec_()
        self.centralWidget().update()

    def add_values_to_scenario(self):
        self.appdata.scen_values_set_multiple(list(self.appdata.project.comp_values.keys()),
                                              list(self.appdata.project.comp_values.values()))
        self.appdata.project.comp_values.clear()
        if self.appdata.auto_fba:
            self.fba()
        self.centralWidget().update()

    def set_model_bounds_to_scenario(self):
        for reaction in self.appdata.project.cobra_py_model.reactions:
            if reaction.id in self.appdata.project.scen_values:
                (vl, vu) = self.appdata.project.scen_values[reaction.id]
                reaction.lower_bound = vl
                reaction.upper_bound = vu
        self.centralWidget().update()

    @Slot()
    def pin_scenario_reactions(self):
        self.centralWidget().reaction_list.pin_multiple(self.appdata.project.scen_values.keys())

    def set_scenario_to_default_scenario(self):
        ''' write current scenario into sbml annotation '''
        for reaction in self.appdata.project.cobra_py_model.reactions:
            if reaction.id in self.appdata.project.scen_values:
                values = self.appdata.project.scen_values[reaction.id]
                reaction.annotation['cnapy-default'] = str(values)
            else:
                if 'cnapy-default' in reaction.annotation.keys():
                    reaction.annotation.pop('cnapy-default')
        self.centralWidget().update()
        self.unsaved_changes()

    def auto_fba(self):
        if self.auto_fba_action.isChecked():
            self.appdata.auto_fba = True
            self.fba()
        else:
            self.appdata.auto_fba = False

    def fba(self):
        with self.appdata.project.cobra_py_model as model:
            self.appdata.project.load_scenario_into_model(model)
            self.appdata.project.solution = model_optimization_with_exceptions(model)
        self.process_fba_solution()

    def process_fba_solution(self, update=True):
        general_solution_error = True
        if hasattr(self.appdata.project, "solution"):
            if hasattr(self.appdata.project.solution, "status"):
                general_solution_error = False
                if self.appdata.project.solution.status == 'optimal':
                    display_text = "Optimal solution with objective value "+self.appdata.format_flux_value(self.appdata.project.solution.objective_value)
                    self.set_status_optimal()
                    for r, v in self.appdata.project.solution.fluxes.items():
                        self.appdata.project.comp_values[r] = (v, v)
                elif self.appdata.project.solution.status == 'infeasible':
                    display_text = "No solution, the current scenario is infeasible"
                    self.set_status_infeasible()
                    self.appdata.project.comp_values.clear()
                else:
                    display_text = "No optimal solution, solver status is "+self.appdata.project.solution.status
                    self.set_status_unknown()
                    self.appdata.project.comp_values.clear()

        if not general_solution_error:
            self.centralWidget().console._append_plain_text("\n"+display_text, before_prompt=True)
            self.solver_status_display.setText(display_text)
            self.appdata.project.comp_values_type = 0
        if update:
            self.centralWidget().update()

    def make_scenario_feasible(self):
        if self.make_scenario_feasible_dialog is None:
            self.make_scenario_feasible_dialog = FluxFeasibilityDialog(self)
        else:
            self.make_scenario_feasible_dialog.modified_scenario = None
        self.make_scenario_feasible_dialog.show()

    def fba_optimize_reaction(self, reaction: str, mmin: bool):
        with self.appdata.project.cobra_py_model as model:
            self.appdata.project.load_scenario_into_model(model)
            model.objective = model.reactions.get_by_id(reaction)
            if mmin:
                model.objective.direction = 'min'
            else:
                model.objective.direction = 'max'
            self.appdata.project.solution = model_optimization_with_exceptions(model)
        self.process_fba_solution()

    def pfba(self):
        with self.appdata.project.cobra_py_model as model:
            self.appdata.project.load_scenario_into_model(model)
            try:
                solution = cobra.flux_analysis.pfba(model)
            except cobra.exceptions.Infeasible:
                display_text = "No solution, the current scenario is infeasible"
                self.set_status_infeasible()
                self.appdata.project.comp_values.clear()
            except Exception:
                display_text = "An unexpected error occured."
                self.set_status_unknown()
                self.appdata.project.comp_values.clear()
                exstr = get_last_exception_string()
                # Check for substrings of Gurobi and CPLEX community edition errors
                if has_community_error_substring(exstr):
                    except_likely_community_model_error()
                else:
                    print(exstr)
                    utils.show_unknown_error_box(exstr)
            else:
                if solution.status == 'optimal':
                    soldict = solution.fluxes.to_dict()
                    for i in soldict:
                        self.appdata.project.comp_values[i] = (
                            soldict[i], soldict[i])
                    display_text = "Optimal solution with objective value "+ \
                        self.appdata.format_flux_value(solution.objective_value)
                    self.set_status_optimal()
                else:
                    display_text = "No optimal solution, solver status is "+solution.status
                    self.set_status_unknown()
                    self.appdata.project.comp_values.clear()
            finally:
                self.centralWidget().console._append_plain_text("\n"+display_text, before_prompt=True)
                self.solver_status_display.setText(display_text)
                self.appdata.project.comp_values_type = 0
                self.centralWidget().update()

    def execute_print_model_stats(self):
        if len(self.appdata.project.cobra_py_model.reactions) > 0:
            self.centralWidget().kernel_client.execute("cna.print_model_stats()")
        else:
            self.centralWidget().kernel_client.execute("print('\\nEmpty matrix!')")

        self.centralWidget().show_bottom_of_console()

    def show_net_conversion(self):
        self.centralWidget().kernel_client.execute("cna.net_conversion()")
        self.centralWidget().show_bottom_of_console()

    def net_conversion(self):
        with self.appdata.project.cobra_py_model as model:
            self.appdata.project.load_scenario_into_model(model)
            solution = model_optimization_with_exceptions(model)
            if solution.status == 'optimal':
                errors = False
                imports = []
                exports = []
                soldict = solution.fluxes.to_dict()
                for i in soldict:
                    r = self.appdata.project.cobra_py_model.reactions.get_by_id(
                        i)
                    val = round(soldict[i], self.appdata.rounding)
                    if r.reactants == []:
                        if len(r.products) != 1:
                            print(
                                'Error: Expected only import reactions with one metabolite but',
                                i, 'imports', r.products)
                            errors = True
                        else:
                            if val > 0.0:
                                imports.append(
                                    str(val) + ' ' + r.products[0].id)
                            elif val < 0.0:
                                exports.append(
                                    str(abs(val)) + ' ' + r.products[0].id)

                    elif r.products == []:
                        if len(r.reactants) != 1:
                            print(
                                'Error: Expected only export reactions with one metabolite but',
                                i, 'exports', r.reactants)
                            errors = True
                        else:
                            if val > 0.0:
                                exports.append(
                                    str(val) + ' ' + r.reactants[0].id)
                            elif val < 0.0:
                                imports.append(
                                    str(abs(val)) + ' ' + r.reactants[0].id)

                if errors:
                    return

                print(
                    '\n\x1b[1;04;30m'+"Net conversion of external metabolites by the given scenario is:\x1b[0m\n")
                print(' + '.join(imports))
                print('-->')
                print(' + '.join(exports))

            elif solution.status == 'infeasible':
                print('No solution the scenario is infeasible!')
            else:
                print('No solution!', solution.status)

    def print_model_stats(self):
        m = cobra.util.array.create_stoichiometric_matrix(
            self.appdata.project.cobra_py_model, array_type='DataFrame')
        metabolites = m.shape[0]
        reactions = m.shape[1]
        print('Stoichiometric matrix:\n', m)
        print('\nNumber of metabolites: ', metabolites)
        print('Number of reactions: ', reactions)
        rank = np.linalg.matrix_rank(m)
        print('\nRank of stoichiometric matrix: ' + str(rank))
        print('Degrees of freedom: ' + str(reactions-rank))
        print('Independent conservation relations: ' + str(metabolites-rank))

        has_non_zero = False
        mmin = None
        abs_m = np.absolute(m.to_numpy())
        for r in abs_m:
            for e in r:
                if not has_non_zero:
                    if e > 0.0:
                        has_non_zero = True
                        mmin = e
                else:
                    if e > 0.0 and e < mmin:
                        mmin = e
        if has_non_zero:
            print('\nSmallest (absolute) non-zero-value:', mmin)
        else:
            print('\nIt\'s the zero matrix')

        c = []
        abs_m = np.absolute(m.to_numpy())
        for r in abs_m:
            x = max(r)
            c.append(x)
        print('Largest (absolute) value:', max(c))

    def print_in_out_fluxes(self, metabolite):
        soldict = {id: val[0] for (id, val) in self.appdata.project.comp_values.items()}
        self.in_out_fluxes(metabolite, soldict)

    def show_model_bounds(self):
        for reaction in self.appdata.project.cobra_py_model.reactions:
            self.appdata.project.comp_values[reaction.id] = (
                reaction.lower_bound, reaction.upper_bound)
        self.appdata.project.comp_values_type = 1
        self.centralWidget().update()

    def fva(self, fraction_of_optimum=0.0, zero_objective_with_zero_fraction_of_optimum=True):
        self.setCursor(Qt.BusyCursor)
        with self.appdata.project.cobra_py_model as model:
            self.appdata.project.load_scenario_into_model(model)
            if zero_objective_with_zero_fraction_of_optimum:
                # completely remove objective for basic FVA, not the same as only setting fraction_of_optimum = 0.0
                model.objective = model.problem.Objective(Zero)
            if len(self.appdata.project.scen_values) > 0 or len(self.appdata.project.scen_values.reactions) > 0:
                update_stoichiometry_hash = True
            else:
                update_stoichiometry_hash = False
            for r in self.appdata.project.cobra_py_model.reactions:
                if r.lower_bound == -float('inf'):
                    r.lower_bound = cobra.Configuration().lower_bound
                    r.set_hash_value()
                    update_stoichiometry_hash = True
                if r.upper_bound == float('inf'):
                    r.upper_bound = cobra.Configuration().upper_bound
                    r.set_hash_value()
                    update_stoichiometry_hash = True
            if self.appdata.use_results_cache:
                if update_stoichiometry_hash:
                    model.set_stoichiometry_hash_object()
                fva_hash = model.stoichiometry_hash_object.copy()
                if len(self.appdata.project.scen_values.constraints) > 0:
                    # although the constraints are already in the model they are not covered by
                    # the reaction hashes and therefore taken into account here
                    fva_hash.update(pickle.dumps(sorted(self.appdata.project.scen_values.constraints)))
            else:
                fva_hash = None
            try:
                solution = flux_variability_analysis(model, fraction_of_optimum=fraction_of_optimum,
                    results_cache_dir=self.appdata.results_cache_dir if self.appdata.use_results_cache else None,
                    fva_hash= fva_hash,
                    print_func=lambda *txt: self.statusBar().showMessage(' '.join(list(txt))))
            except cobra.exceptions.Infeasible:
                QMessageBox.information(
                    self, 'No solution', 'The scenario is infeasible')
            except Exception:
                exstr = get_last_exception_string()
                # Check for substrings of Gurobi and CPLEX community edition errors
                if has_community_error_substring(exstr):
                    except_likely_community_model_error()
                else:
                    print(exstr)
                    utils.show_unknown_error_box(exstr)
            else:
                minimum = solution.minimum.to_dict()
                maximum = solution.maximum.to_dict()
                for i in minimum:
                    self.appdata.project.comp_values[i] = (
                        minimum[i], maximum[i])
                self.appdata.project.fva_values = self.appdata.project.comp_values.copy()
                self.appdata.project.comp_values_type = 1

        self.centralWidget().update()
        self.setCursor(Qt.ArrowCursor)

    # def efm(self):
    #     self.efm_dialog = EFMDialog(
    #         self.appdata, self.centralWidget())
    #     self.efm_dialog.exec_()

    def in_out_flux(self):
        in_out_flux_dialog = InOutFluxDialog(
            self.appdata)
        in_out_flux_dialog.exec_()

    def all_in_out_fluxes(self):
        filename = self._get_filename("xlsx")
        if filename is None:
            return

        soldict = {
            id: val[0] for (id, val) in self.appdata.project.comp_values.items()
        }

        fluxes_per_metabolite = {}
        with self.appdata.project.cobra_py_model as model:
            self.appdata.project.scen_values.add_scenario_reactions_to_model(model)
            for metabolite in model.metabolites:
                for reaction in metabolite.reactions:
                    if reaction.id not in soldict.keys():
                        continue
                    if abs(soldict[reaction.id]) < self.appdata.project.cobra_py_model.tolerance:
                        continue

                    if (metabolite.id, metabolite.name) not in fluxes_per_metabolite.keys():
                        fluxes_per_metabolite[(metabolite.id, metabolite.name)] = []

                    stoichiometry = reaction.metabolites[metabolite]
                    fluxes_per_metabolite[(metabolite.id, metabolite.name)].append([
                        stoichiometry * soldict[reaction.id],
                        reaction.id,
                        reaction.reaction,
                    ])

        # Sheet styles
        italic = openpyxl.styles.Font(italic=True)
        bold = openpyxl.styles.Font(bold=True)
        underlined = openpyxl.styles.Font(underline="single")

        # Main spreadsheet variable
        wb = openpyxl.Workbook()

        # In and out sums sheet
        ws1 = wb.create_sheet("In and out sums")
        cell = ws1.cell(1, 1)
        cell.value = "Metabolite ID"
        cell.font = bold

        cell = ws1.cell(1, 2)
        cell.value = "Metabolite name"
        cell.font = bold

        cell = ws1.cell(1, 3)
        cell.value = "In flux sum"
        cell.font = bold

        cell = ws1.cell(1, 4)
        cell.value = "Out flux sum"
        cell.font = bold

        current_line = 2
        for met_data, reac_data in fluxes_per_metabolite.items():
            positive_fluxes = [x[0] for x in reac_data if x[0] > 0.0]
            negative_fluxes = [x[0] for x in reac_data if x[0] < 0.0]

            cell = ws1.cell(current_line, 1)
            cell.value = met_data[0]
            cell = ws1.cell(current_line, 2)
            cell.value = met_data[1]
            cell = ws1.cell(current_line, 3)
            cell.value = sum(positive_fluxes)
            cell = ws1.cell(current_line, 4)
            cell.value = sum(negative_fluxes)

            current_line += 1

        ws1.column_dimensions['A'].width = 16
        ws1.column_dimensions['B'].width = 18
        ws1.column_dimensions['C'].width = 16
        ws1.column_dimensions['D'].width = 16

        # Details sheet
        ws2 = wb.create_sheet("Details")
        current_line = 1
        for met_data, reac_data in fluxes_per_metabolite.items():
            positive_reactions = [x for x in reac_data if x[0] > 0.0]
            positive_reactions = sorted(positive_reactions, key=lambda x: x[1])
            negative_reactions = [x for x in reac_data if x[0] < 0.0]
            negative_reactions = sorted(negative_reactions, key=lambda x: x[1])

            cell = ws2.cell(current_line, 1)
            cell.value = "Metabolite ID:"
            cell.font = bold
            cell = ws2.cell(current_line, 2)
            cell.value = met_data[0]

            for reaction_set in (("Producing", positive_reactions), ("Consuming", negative_reactions)):
                current_line += 1
                cell = ws2.cell(current_line, 1)
                cell.value = f"{reaction_set[0]} reactions:"
                cell.font = underlined

                current_line += 1
                cell = ws2.cell(current_line, 1)
                cell.value = f"{reaction_set[0]} flux"
                cell.font = italic

                cell = ws2.cell(current_line, 2)
                cell.value = "Reaction ID"
                cell.font = italic

                cell = ws2.cell(current_line, 3)
                cell.value = "Reaction string"
                cell.font = italic

                for reaction_data in reaction_set[1]:
                    current_line += 1
                    cell = ws2.cell(current_line, 1)
                    cell.value = reaction_data[0]
                    cell = ws2.cell(current_line, 2)
                    cell.value = reaction_data[1]
                    cell = ws2.cell(current_line, 3)
                    cell.value = reaction_data[2]
            current_line += 2

        ws2.column_dimensions['A'].width = 20
        ws2.column_dimensions['B'].width = 16
        ws2.column_dimensions['C'].width = 16

        del(wb["Sheet"])
        wb.save(filename)


    def efmtool(self):
        self.efmtool_dialog = EFMtoolDialog(
            self.appdata, self.centralWidget())
        self.efmtool_dialog.exec_()

    def mcs(self):
        if self.mcs_dialog is None:
            self.mcs_dialog = MCSDialog(self.appdata, self.centralWidget())
        self.mcs_dialog.show()

    def set_onoff(self):
        self.centralWidget().set_onoff()

    def set_heaton(self):
        self.centralWidget().set_heaton()

    def in_out_fluxes(self, metabolite_id, soldict):
        self.centralWidget().kernel_client.execute('%matplotlib inline', store_history=False)
        with self.appdata.project.cobra_py_model as model:
            self.appdata.project.scen_values.add_scenario_reactions_to_model(model)
            met = model.metabolites.get_by_id(metabolite_id)
            fig, ax = plt.subplots()
            ax.set_xticks([1, 2])
            ax.set_xticklabels(['In', 'Out'])
            cons = []
            prod = []
            sum_cons = 0
            sum_prod = 0
            for rxn in met.reactions:
                flux = soldict.get(rxn.id, 0.0)
                if abs(flux) > model.tolerance:
                    flux *= rxn.get_coefficient(metabolite_id)
                    if flux < 0:
                        cons.append((rxn, -flux))
                    elif flux > 0:
                        prod.append((rxn, flux))
            cons = sorted(cons, key=lambda x: x[1], reverse=True)
            prod = sorted(prod, key=lambda x: x[1], reverse=True)
            for rxn, flux in prod:
                ax.bar(1, flux, width=0.8, bottom=sum_prod, label=rxn.id+": "+rxn.build_reaction_string())
                sum_prod += flux
            for rxn, flux in cons:
                ax.bar(2, flux, width=0.8, bottom=sum_cons, label=rxn.id+": "+rxn.build_reaction_string())
                sum_cons += flux
            ax.set_ylabel('Flux')
            ax.set_title('In/Out fluxes at metabolite ' + metabolite_id)
            ax.legend(bbox_to_anchor=(1, 1), loc="upper left")

            # Print plot in CNApy's console
            plt.show()

            # Pretty print cons and prod lists of tuples
            pretty_prod_dict = f"\nProducing reactions of {metabolite_id}:\n"+json.dumps({
                x[0].id: x[1]
                for x in prod
            }, indent=2)
            pretty_cons_dict = f"\nConsuming reactions of {metabolite_id}:\n"+json.dumps({
                x[0].id: x[1]
                for x in cons
            }, indent=2)
            # The next print statements are directly executed in CNApy's Jupyter console
            print(pretty_prod_dict)
            print(pretty_cons_dict)

        self.centralWidget().kernel_client.execute('%matplotlib qt', store_history=False)

        return prod, cons

    def show_console(self):
        print("show model view")
        (x, _) = self.centralWidget().splitter.sizes()
        if x < 50:
            self.show_model_view()
        (_, r) = self.centralWidget().splitter2.getRange(1)
        self.centralWidget().splitter2.moveSplitter(round(r*0.5), 1)

    def show_map_view(self):
        self.show_console()

    def show_model_view(self):
        (_, r) = self.centralWidget().splitter.getRange(1)
        self.centralWidget().splitter.moveSplitter(round(r*0.5), 1)

    def clear_status_bar(self):
        self.solver_status_display.setText("")
        self.solver_status_symbol.setText("")

    def set_status_optimal(self):
        self.solver_status_symbol.setStyleSheet("color: green; font-weight: bold")
        self.solver_status_symbol.setText("\u2713")

    def set_status_infeasible(self):
        self.solver_status_symbol.setStyleSheet("color: red; font-weight: bold")
        self.solver_status_symbol.setText("\u2717")

    def set_status_unknown(self):
        self.solver_status_symbol.setStyleSheet("color: black")
        self.solver_status_symbol.setText("?")

    @Slot()
    def perform_optmdfpathway(self):
        # Has to be in self to keep computation thread
        self.optmdfpathway_dialog = ThermodynamicDialog(
            self.appdata,
            self.centralWidget(),
            analysis_type=ThermodynamicAnalysisTypes.OPTMDFPATHWAY
        )
        self.optmdfpathway_dialog.exec_()

    @Slot()
    def perform_thermodynamic_fba(self):
        # Has to be in self to keep computation thread
        self.thermodynamic_fba_dialog = ThermodynamicDialog(
            self.appdata,
            self.centralWidget(),
            analysis_type=ThermodynamicAnalysisTypes.THERMODYNAMIC_FBA
        )
        self.thermodynamic_fba_dialog.exec_()

    @Slot()
    def perform_bottleneck_analysis(self):
        # Has to be in self to keep computation thread
        self.bottleneck_dialog = ThermodynamicDialog(
            self.appdata,
            self.centralWidget(),
            analysis_type=ThermodynamicAnalysisTypes.BOTTLENECK_ANALYSIS
        )
        self.bottleneck_dialog.exec_()

    def _load_json(self) -> Dict[Any, Any]:
        dialog = QFileDialog(self)
        filename: str = dialog.getOpenFileName(
            directory=self.appdata.last_scen_directory, filter="*.json")[0]
        if not filename or len(filename) == 0 or not os.path.exists(filename):
            return {}

        try:
            with open(filename) as f:
                json_data = json.load(f)
        except json.decoder.JSONDecodeError:
            QMessageBox.critical(
                self,
                "Could not open file",
                "File could not be opened as it does not seem to be a valid JSON file."
            )
            return {}

        return json_data

    def _load_active_xlsx_worksheet(self):
        dialog = QFileDialog(self)
        filename: str = dialog.getOpenFileName(
            directory=self.appdata.last_scen_directory, filter="*.xlsx")[0]
        if not filename or len(filename) == 0 or not os.path.exists(filename):
            return
        wb = openpyxl.load_workbook(filename, read_only=True)
        ws = wb.active
        return ws

    def _set_concentrations(self, concentrations, replace_all):
        for metabolite in self.appdata.project.cobra_py_model.metabolites:
            if replace_all:
                if "Cmin" in metabolite.annotation:
                    del(metabolite.annotation["Cmin"])
                if "Cmax" in metabolite.annotation:
                    del(metabolite.annotation["Cmax"])
            if metabolite.id not in concentrations.keys():
                if "DEFAULT" in concentrations.keys():
                    lb = concentrations["DEFAULT"]["min"]
                    ub = concentrations["DEFAULT"]["max"]
                else:
                    continue
            else:
                lb = concentrations[metabolite.id]["min"]
                ub = concentrations[metabolite.id]["max"]

            metabolite.annotation["Cmin"] = lb
            metabolite.annotation["Cmax"] = ub
        self.centralWidget().update()
        self.unsaved_changes()

    def _set_dG0s(self, dG0s, replace_all):
        if replace_all:
            for reaction in self.appdata.project.cobra_py_model.reactions:
                if "dG0" in reaction.annotation:
                    del(reaction.annotation["dG0"])
                if "dG0_uncertainty" in reaction.annotation:
                    del(reaction.annotation["dG0_uncertainty"])

        reaction_ids = [x.id for x in self.appdata.project.cobra_py_model.reactions]
        for reaction_id in dG0s.keys():
            if reaction_id not in reaction_ids:
                continue
            reaction = self.appdata.project.cobra_py_model.reactions.get_by_id(reaction_id)
            reaction.annotation["dG0"] = dG0s[reaction_id]["dG0"]
            if "uncertainty" in dG0s[reaction_id].keys():
                reaction.annotation["dG0_uncertainty"] = dG0s[reaction_id]["uncertainty"]
        self.centralWidget().update()
        self.unsaved_changes()

    def _load_concentrations_json(self, replace_all):
        concentrations = self._load_json()
        self._set_concentrations(concentrations)

    def load_concentrations_json_amend(self, replace_all):
        self._load_concentrations_json(False)

    def load_concentrations_json_replace_all(self, replace_all):
        self._load_concentrations_json(True)

    def _load_concentrations_xlsx(self, replace_all):
        ws = self._load_active_xlsx_worksheet()

        if ws is None:
            return

        metabolite_id_column = 1
        cmin_column = 2
        cmax_column = 3

        concentrations: Dict[str, Dict[str, float]] = {}
        warnings = ""
        for row in range(2, ws.max_row+1):
            metabolite_id = ws.cell(row=row, column=metabolite_id_column).value
            cmin_in_cell = ws.cell(row=row, column=cmin_column).value
            cmax_in_cell = ws.cell(row=row, column=cmax_column).value

            if metabolite_id is None:
                continue

            if cmin_in_cell is not None:
                try:
                    cmin = float(cmin_in_cell)
                except ValueError:
                    warnings += f"WARNING: Cmin of {metabolite_id} could not be read as number. "\
                                "This metabolite will be ignored.\n"
                    continue
            else:
                warnings += f"WARNING: No Cmin for {metabolite_id}. "\
                            "This metabolite will be ignored.\n"
                continue

            if cmax_in_cell is not None:
                try:
                    cmax = float(cmax_in_cell)
                except ValueError:
                    warnings += f"WARNING: Cmin of {metabolite_id} could not be read as number. "\
                                "This metabolite will be ignored.\n"
                    continue
            else:
                warnings += f"WARNING: No Cmax for {metabolite_id}. "\
                                "This metabolite will be ignored.\n"
                continue
            print(warnings)
            concentrations[metabolite_id] = {
                "min": cmin,
                "max": cmax,
            }
        if warnings != "":
            QMessageBox.warning(
                None,
                "Warnings occured while loading XLSX",
                f"The following warnings occured while loading the XLSX:\n{warnings}"
            )
        self._set_concentrations(concentrations)

    def load_concentrations_xlsx_amend(self):
        self._load_concentrations_xlsx(False)

    def load_concentrations_xlsx_replace_all(self):
        self._load_concentrations_xlsx(True)

    def _load_dG0_json(self, replace_all):
        dG0s = self._load_json()
        self._set_dG0s(dG0s, replace_all)

    def load_dG0_json_amend(self):
        self._load_dG0_json(False)

    def load_dG0_json_replace_all(self):
        self._load_dG0_json(True)

    def _load_dG0_xlsx(self, replace_all):
        ws = self._load_active_xlsx_worksheet()

        if ws is None:
            return

        reac_id_column = 1
        dG0_column = 2
        uncertainty_column = 3

        dG0s: Dict[str, Dict[str, float]] = {}
        warnings = ""
        for row in range(2, ws.max_row+1):
            reac_id = ws.cell(row=row, column=reac_id_column).value
            dG0_in_cell = ws.cell(row=row, column=dG0_column).value
            uncertainty_in_cell = ws.cell(row=row, column=uncertainty_column).value

            if reac_id is None:
                continue

            if dG0_in_cell is not None:
                try:
                    dG0 = float(dG0_in_cell)
                except ValueError:
                    warnings += f"WARNING: dG'° of {reac_id} could not be read as number. "\
                                "It will be ignored.\n"
                    continue

                dG0s[reac_id] = {}
                dG0s[reac_id]["dG0"] = dG0

                if uncertainty_in_cell is not None:
                    try:
                        uncertainty = float(uncertainty_in_cell)
                    except ValueError:
                        warnings += f"WARNING: Uncertainty of {reac_id} could not be read"\
                                    "as number. It will be ignored.\n"
                        continue
                    dG0s[reac_id]["uncertainty"] = uncertainty
            elif uncertainty_in_cell is not None:
                warnings += f"WARNING: Uncertainty of {reac_id} is set but no dG'°"\
                            " value exists. Hence, it will be ignored.\n"
        if warnings != "":
            QMessageBox.warning(
                None,
                "Warnings occured while loading XLSX",
                f"The following warnings occured while loading the XLSX:\n{warnings}"
            )
        self._set_dG0s(dG0s, replace_all)

    def load_dG0_xlsx_amend(self):
        self._load_dG0_xlsx(False)

    def load_dG0_xlsx_replace_all(self):
        self._load_dG0_xlsx(True)

    def _get_filename(self, filetype: str) -> str:
        dialog = QFileDialog(self)
        filename: str = dialog.getSaveFileName(
            directory=self.appdata.work_directory, filter=f"*.{filetype}")[0]
        if not filename or len(filename) == 0:
            return
        if not (filename.endswith(f".{filetype}")):
            filename += f".{filetype}"
        return filename


    def _save_fluxes(self, filetype: str):
        filename = self._get_filename(filetype)

        table = self.central_widget.reaction_list.get_as_table()

        if filetype == "csv":
            with open(filename, "w", encoding="utf-8") as f:
                f.write(table)
        elif filetype == "xlsx":
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "CNApy"
            lines = table.split("\r")
            current_line = 1
            for line in lines:
                current_column = 1
                cells = line.split("\t")
                for cell in cells:
                    sheet_cell = sheet.cell(row=current_line, column=current_column)
                    sheet_cell.value = cell
                    current_column += 1
                current_line += 1
            wb.save(filename)

    def save_fluxes_as_csv(self):
        self._save_fluxes("csv")

    def save_fluxes_as_xlsx(self):
        self._save_fluxes("xlsx")

    @Slot()
    def update_scenario_file_name(self):
        if len(self.appdata.project.scen_values.file_name) == 0:
            self.load_scenario_action_tb.setIconText("No scenario file loaded")
            self.reload_scenario_action.setEnabled(False)
            self.save_scenario_action.setEnabled(False)
        else:
            dir_name, file_name = os.path.split(
                self.appdata.project.scen_values.file_name)
            if self.appdata.project.scen_values.has_unsaved_changes:
                file_name += "*"
            self.load_scenario_action_tb.setIconText(
                os.path.basename(dir_name) + os.path.sep + file_name)
            self.reload_scenario_action.setEnabled(True)
            self.save_scenario_action.setEnabled(True)

    @Slot()
    def reload_scenario(self):
        self.load_scenario_file(self.appdata.project.scen_values.file_name)
